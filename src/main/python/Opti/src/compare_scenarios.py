#!/usr/bin/env python
from __future__ import annotations

import argparse
import json
import math
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from input_scenarios import resolve_scenario_paths

PROJECT_ROOT = Path(__file__).resolve().parents[1]

TECH_ORDER = {
    "Chargers only": 0,
    "Chargers + PV": 1,
    "Chargers + BESS": 2,
    "Chargers + PV + BESS": 3,
}
REDIR_ORDER = {"No redirection": 0, "With redirection": 1}

SUMMARY_SPEC = [
    ("Charging infrastructure", None, None, "section"),
    ("Home chargers, exogenous (7 kW)", "home_chargers_exogenous", "count", "always"),
    ("Slow chargers (11 kW)", "chargers_slow_installed", "count", "always"),
    ("Medium chargers (22 kW)", "chargers_medium_installed", "count", "always"),
    ("Fast chargers (50 kW)", "chargers_fast_installed", "count", "always"),
    ("Public charger capacity (kWh/30 min)", "public_charger_capacity_kWh_per_slot", "number", "always"),
    ("PV panels (500 W)", "PV_panels_installed", "count", "pv"),
    ("Battery units (10 kWh)", "battery_units_installed", "count", "bess"),

    ("Energy flows (kWh/yr)", None, None, "section"),
    ("Grid electricity purchased", "grid_total_kWh", "number", "always"),
    ("Direct grid supply to chargers", "grid_direct_kWh", "number", "always"),
    ("Grid energy to BESS", "grid_to_battery_kWh", "number", "bess"),
    ("PV electricity used locally", "pv_used_total_kWh", "number", "pv"),
    ("Direct PV supply to chargers", "pv_direct_kWh", "number", "pv"),
    ("PV energy to BESS", "pv_to_battery_kWh", "number", "pv_bess"),
    ("Battery discharge", "battery_discharge_kWh", "number", "bess"),
    ("Redirected public charging demand", "energy_redirected_kWh", "number", "redir"),

    ("Derived demand and flexibility indicators", None, None, "section"),
    ("Estimated redirected charging trip equivalents/yr", "redirection_trip_equivalents_annual", "count", "redir"),
    ("Whole redirected trip bundles/yr", "redirection_full_trip_bundles_annual", "count", "redir"),
    ("Share of public demand redirected", "share_redirected_public_demand", "percent", "redir"),
    ("PV use relative to CPO charging demand", "derived_pv_use_share", "percent", "pv"),
    ("Annual public charging demand", "annual_public_demand_kWh", "number", "always"),
    ("Annual residual home demand at CPO boundary", "annual_home_residual_demand_kWh", "number", "always"),
    ("Annual optimizer-boundary charging demand", "annual_optimizer_boundary_demand_kWh", "number", "always"),
    ("Annual total MATSim charging demand", "annual_total_MATSim_charging_demand_kWh", "number", "always"),
    ("Annual unmet charging demand", "annual_slack_kWh", "number4", "always"),
    ("Unmet demand share of optimizer-boundary demand", "derived_slack_share", "percent6", "always"),

    ("Annual economic terms (SEK/yr)", None, None, "section"),
    ("Public charging revenue", "revenue_all_chargers_SEK", "currency", "always"),
    ("Grid electricity cost", "grid_cost_SEK", "currency", "always"),
    ("User-redirection incentive cost", "redirection_total_cost_SEK", "currency", "redir"),
    ("Distance compensation", "redirection_distance_cost_SEK", "currency", "redir"),
    ("Price compensation", "redirection_price_compensation_SEK", "currency", "redir"),
    ("Unmet demand penalty", "slack_penalty_SEK", "currency", "always"),
    ("Annualized charger CapEx", "capex_chargers_SEK", "currency", "always"),
    ("Annualized PV CapEx", "capex_PV_SEK", "currency", "pv"),
    ("Annualized BESS CapEx", "capex_BESS_SEK", "currency", "bess"),
    ("Annualized PV and BESS CapEx", "capex_PV_BESS_SEK", "currency", "pv_or_bess"),
    ("Net profit", "annual_profit_SEK", "currency", "always"),

    ("Comparative indicators", None, None, "section"),
    ("Net profit gain vs selected baseline", "derived_profit_gain", "currency", "always"),
    ("Net profit gain vs selected baseline (%)", "derived_profit_gain_pct", "percent", "always"),
    ("Grid electricity reduction vs selected baseline", "derived_grid_reduction", "number", "always"),
    ("Grid electricity reduction vs selected baseline (%)", "derived_grid_reduction_pct", "percent", "always"),
    ("Public charger capacity reduction vs selected baseline", "derived_capacity_reduction", "number", "always"),
    ("Net profit gain from user redirection", "derived_redirection_profit_gain", "currency", "redir"),
]


def _finite(value: Any) -> float:
    try:
        number = float(value)
        return number if math.isfinite(number) else math.nan
    except Exception:
        return math.nan


def _load_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _load_summary(run_dir: Path) -> dict[str, Any]:
    path = run_dir / "results" / "model_summary.csv"
    if not path.exists():
        raise FileNotFoundError(f"Missing required result file: {path}")
    frame = pd.read_csv(path)
    if not {"Metric", "Value"}.issubset(frame.columns):
        raise ValueError(f"Expected Metric/Value columns in {path}")
    out: dict[str, Any] = {}
    for row in frame.itertuples(index=False):
        key = str(row.Metric)
        value = row.Value
        try:
            out[key] = float(value)
        except Exception:
            out[key] = str(value)
    return out


def _read_text(path: Path) -> str:
    if not path.exists():
        return ""
    try:
        return path.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return ""


def _find_transcript_value(text: str, key: str) -> str | None:
    match = re.search(rf"(?mi)^\s*{re.escape(key)}\s*:\s*(.+?)\s*$", text)
    return match.group(1).strip() if match else None


def _detect_method(run_dir: Path, metadata: dict[str, Any], transcript: str) -> str:
    method = str(metadata.get("method", "")).strip().lower()
    if method == "lbbd":
        return "LBBD"
    if method == "monolithic":
        return "Monolithic"
    if "LBBD" in run_dir.name.upper() or "LBBD" in transcript[:1000].upper():
        return "LBBD"
    if "MONOLITHIC" in transcript[:1000].upper():
        return "Monolithic"
    return "Unknown"


def _detect_dataset(run_dir: Path, summary: dict[str, Any], metadata: dict[str, Any], transcript: str) -> str:
    for candidate in (
        metadata.get("dataset"),
        summary.get("dataset"),
        _find_transcript_value(transcript, "Dataset"),
    ):
        if candidate is not None and str(candidate).strip().lower() in {"small", "full"}:
            return str(candidate).strip().lower()
    match = re.search(r"(?:^|_)(small|full)(?:_|$)", run_dir.name, flags=re.I)
    return match.group(1).lower() if match else "unknown"


def _detect_vipv(run_dir: Path, summary: dict[str, Any], metadata: dict[str, Any], transcript: str) -> str:
    for candidate in (
        metadata.get("vipv_scenario"),
        summary.get("vipv_scenario"),
        _find_transcript_value(transcript, "VIPV scenario"),
    ):
        if candidate is not None and str(candidate).strip():
            return str(candidate).strip()
    match = re.search(r"(noVIPV|VIPV(?:20|50|80)_Wp(?:400|700|1000))", run_dir.name)
    return match.group(1) if match else "unknown"


def _detect_redirection(run_dir: Path, metadata: dict[str, Any], transcript: str) -> str:
    candidates = [metadata.get("scenario"), _find_transcript_value(transcript, "Scenario")]
    for candidate in candidates:
        text = str(candidate or "").strip().lower()
        if text == "with_redirection":
            return "With redirection"
        if text == "no_redirection":
            return "No redirection"
    name = run_dir.name.lower()
    if "_with_redirection_" in name:
        return "With redirection"
    if "_no_redirection_" in name:
        return "No redirection"
    return "Unknown"


def _bool_from_text(value: str | None) -> bool | None:
    if value is None:
        return None
    text = value.strip().lower()
    if text in {"true", "1", "yes"}:
        return True
    if text in {"false", "0", "no"}:
        return False
    return None


def _detect_technology(run_dir: Path, metadata: dict[str, Any], transcript: str) -> tuple[str, bool, bool]:
    name = run_dir.name
    lower = name.lower()
    pv_enabled: bool | None = None
    bess_enabled: bool | None = None

    technology_line = _find_transcript_value(transcript, "Technology")
    if technology_line:
        low = technology_line.lower()
        if "withpv" in low or "pv enabled" in low:
            pv_enabled = True
        if "nopv" in low or "pv disabled" in low:
            pv_enabled = False
        if "withbess" in low or "bess enabled" in low:
            bess_enabled = True
        if "nobess" in low or "bess disabled" in low:
            bess_enabled = False

    disable_pv = _bool_from_text(_find_transcript_value(transcript, "Disable PV"))
    disable_bess = _bool_from_text(_find_transcript_value(transcript, "Disable BESS"))
    if disable_pv is not None:
        pv_enabled = not disable_pv
    if disable_bess is not None:
        bess_enabled = not disable_bess

    settings = metadata.get("effective_settings", {}) if isinstance(metadata.get("effective_settings"), dict) else {}
    if "disable_pv" in settings:
        pv_enabled = not bool(settings["disable_pv"])
    if "disable_bess" in settings:
        bess_enabled = not bool(settings["disable_bess"])

    if "nopv_nobess" in lower:
        pv_enabled, bess_enabled = False, False
    elif "withpv_nobess" in lower:
        pv_enabled, bess_enabled = True, False
    elif "nopv_withbess" in lower:
        pv_enabled, bess_enabled = False, True
    elif "withpv_withbess" in lower:
        pv_enabled, bess_enabled = True, True

    if pv_enabled is None:
        pv_enabled = True
    if bess_enabled is None:
        bess_enabled = True

    if pv_enabled and bess_enabled:
        label = "Chargers + PV + BESS"
    elif pv_enabled:
        label = "Chargers + PV"
    elif bess_enabled:
        label = "Chargers + BESS"
    else:
        label = "Chargers only"
    return label, pv_enabled, bess_enabled


def _read_complexity_scalars(run_dir: Path) -> dict[str, float]:
    path = run_dir / "results" / "computational_complexity_scalars.csv"
    if not path.exists():
        return {}
    try:
        frame = pd.read_csv(path)
    except Exception:
        return {}
    if not {"Metric", "Value"}.issubset(frame.columns):
        return {}
    out = {}
    for row in frame.itertuples(index=False):
        val = _finite(row.Value)
        if math.isfinite(val):
            out[str(row.Metric)] = val
    return out


def _parse_monolithic_final_gap_pct(run_dir: Path, transcript: str) -> float:
    texts = [transcript, _read_text(run_dir / "logs" / "gurobi_run.log")]
    pattern = re.compile(r"Best objective[^\n]*?gap\s+([0-9.eE+\-]+)%", flags=re.I)
    for text in texts:
        matches = pattern.findall(text)
        if matches:
            return _finite(matches[-1])
    return math.nan


def _solver_metrics(run_dir: Path, method: str, metadata: dict[str, Any], transcript: str) -> dict[str, Any]:
    scalars = _read_complexity_scalars(run_dir)
    result: dict[str, Any] = {
        "requested_gap_pct": math.nan,
        "achieved_gap_pct": math.nan,
        "total_runtime_s": scalars.get("total_runtime_seconds", math.nan),
        "solve_seconds": scalars.get("solve_seconds", scalars.get("decomposition_solve_seconds", math.nan)),
        "peak_rss_MB": scalars.get("peak_process_tree_rss_MB", math.nan),
        "threads": math.nan,
        "termination": metadata.get("termination", ""),
    }
    if method == "Monolithic":
        req = scalars.get("solver_mip_gap_requested", math.nan)
        result["requested_gap_pct"] = 100.0 * req if math.isfinite(req) else math.nan
        result["achieved_gap_pct"] = _parse_monolithic_final_gap_pct(run_dir, transcript)
        thread_match = re.findall(r"Thread count was\s+(\d+)", transcript)
        if thread_match:
            result["threads"] = float(thread_match[-1])
        if not result["termination"]:
            term = re.findall(r"Termination condition:\s*([^\n]+)", transcript, flags=re.I)
            result["termination"] = term[-1].strip() if term else ""
    elif method == "LBBD":
        gap = _finite(metadata.get("certified_gap"))
        if math.isfinite(gap):
            result["achieved_gap_pct"] = 100.0 * gap
        settings = metadata.get("effective_settings", {}) if isinstance(metadata.get("effective_settings"), dict) else {}
        target = _finite(settings.get("lbbd_gap"))
        if math.isfinite(target):
            result["requested_gap_pct"] = 100.0 * target
        threads = _finite(settings.get("threads"))
        if math.isfinite(threads):
            result["threads"] = threads
        if not math.isfinite(_finite(result["total_runtime_s"])):
            result["total_runtime_s"] = _finite(metadata.get("elapsed_seconds"))
    return result


def _recompute_home_chargers(project_root: Path, dataset: str, vipv_scenario: str) -> float:
    import geopandas as gpd

    paths = resolve_scenario_paths(project_root, dataset, vipv_scenario)
    cfg = json.loads((project_root / "config" / "model_config.json").read_text(encoding="utf-8"))
    demand = gpd.read_file(paths["demand_shapefile"], columns=["HexID"])
    parking = gpd.read_file(paths["parking_shapefile"])
    if "HexID" not in demand.columns or "HexID" not in parking.columns or "homeChar" not in parking.columns:
        raise ValueError("Cannot reconstruct exogenous home chargers: required HexID/homeChar columns are absent.")
    active_hex = set(pd.to_numeric(demand["HexID"], errors="coerce").dropna().astype(int))
    parking = parking.copy()
    parking["HexID"] = pd.to_numeric(parking["HexID"], errors="coerce")
    parking = parking[parking["HexID"].isin(active_hex)]
    home = pd.to_numeric(parking["homeChar"], errors="coerce").fillna(0.0) + float(cfg.get("home_charger_add", 0))
    return float(home.sum())


def _capacity_from_config(project_root: Path, metrics: dict[str, Any]) -> float:
    cfg = json.loads((project_root / "config" / "model_config.json").read_text(encoding="utf-8"))
    caps = cfg["charger_capacity_kwh_per_slot"]
    return sum(
        _finite(metrics.get(f"chargers_{c}_installed", 0.0)) * float(caps[c])
        for c in ("slow", "medium", "fast")
    )


def _infer_missing_metrics(project_root: Path, run: dict[str, Any], home_cache: dict[tuple[str, str], float], skip_input_recompute: bool) -> None:
    metrics = run["metrics"]
    if not math.isfinite(_finite(metrics.get("public_charger_capacity_kWh_per_slot"))):
        metrics["public_charger_capacity_kWh_per_slot"] = _capacity_from_config(project_root, metrics)

    if not math.isfinite(_finite(metrics.get("home_chargers_exogenous"))) and not skip_input_recompute:
        key = (run["dataset"], run["vipv_scenario"])
        if key not in home_cache and "unknown" not in key:
            home_cache[key] = _recompute_home_chargers(project_root, *key)
        if key in home_cache:
            metrics["home_chargers_exogenous"] = home_cache[key]

    served = sum(
        max(0.0, _finite(metrics.get(key, 0.0)))
        for key in ("grid_direct_kWh", "pv_direct_kWh", "battery_discharge_kWh")
        if math.isfinite(_finite(metrics.get(key, 0.0)))
    )
    pv_used = _finite(metrics.get("pv_used_total_kWh"))
    metrics["derived_pv_use_share"] = pv_used / served if served > 1e-12 and math.isfinite(pv_used) else math.nan

    boundary = _finite(metrics.get("annual_optimizer_boundary_demand_kWh"))
    if not math.isfinite(boundary):
        public = _finite(metrics.get("annual_public_demand_kWh"))
        residual = _finite(metrics.get("annual_home_residual_demand_kWh"))
        if math.isfinite(public) and math.isfinite(residual):
            boundary = public + residual
            metrics["annual_optimizer_boundary_demand_kWh"] = boundary
    slack = _finite(metrics.get("annual_slack_kWh"))
    metrics["derived_slack_share"] = slack / boundary if math.isfinite(slack) and math.isfinite(boundary) and boundary > 0 else math.nan


def _is_applicable(run: dict[str, Any], rule: str) -> bool:
    if rule == "always":
        return True
    if rule == "pv":
        return bool(run["pv_enabled"])
    if rule == "bess":
        return bool(run["bess_enabled"])
    if rule == "pv_bess":
        return bool(run["pv_enabled"] and run["bess_enabled"])
    if rule == "pv_or_bess":
        return bool(run["pv_enabled"] or run["bess_enabled"])
    if rule == "redir":
        return run["redirection"] == "With redirection"
    return True


def _difference(a: Any, b: Any) -> float:
    aa, bb = _finite(a), _finite(b)
    return aa - bb if math.isfinite(aa) and math.isfinite(bb) else math.nan


def _relative_diff(a: Any, b: Any) -> float:
    diff = _difference(a, b)
    bb = _finite(b)
    return diff / abs(bb) if math.isfinite(diff) and math.isfinite(bb) and abs(bb) > 1e-12 else math.nan


def _build_redirection_pair_map(runs: list[dict[str, Any]]) -> dict[int, dict[str, Any]]:
    grouped: dict[tuple[str, str, str], dict[str, list[tuple[int, dict[str, Any]]]]] = defaultdict(lambda: defaultdict(list))
    for idx, run in enumerate(runs):
        key = (run["dataset"], run["vipv_scenario"], run["technology"])
        grouped[key][run["redirection"]].append((idx, run))

    pair_map: dict[int, dict[str, Any]] = {}
    for key, groups in grouped.items():
        no_runs = groups.get("No redirection", [])
        with_runs = groups.get("With redirection", [])
        if len(no_runs) == 1 and len(with_runs) == 1:
            no_idx, no_run = no_runs[0]
            with_idx, with_run = with_runs[0]
            pair_map[with_idx] = {
                "no_idx": no_idx,
                "no_run": no_run,
                "with_run": with_run,
                "key": key,
            }
    return pair_map


def _add_derived_comparisons(runs: list[dict[str, Any]], baseline_idx: int) -> dict[int, dict[str, Any]]:
    baseline = runs[baseline_idx]["metrics"]
    pair_map = _build_redirection_pair_map(runs)
    for idx, run in enumerate(runs):
        metrics = run["metrics"]
        metrics["derived_profit_gain"] = _difference(metrics.get("annual_profit_SEK"), baseline.get("annual_profit_SEK"))
        metrics["derived_profit_gain_pct"] = _relative_diff(metrics.get("annual_profit_SEK"), baseline.get("annual_profit_SEK"))
        metrics["derived_grid_reduction"] = _difference(baseline.get("grid_total_kWh"), metrics.get("grid_total_kWh"))
        base_grid = _finite(baseline.get("grid_total_kWh"))
        grid_reduction = _finite(metrics.get("derived_grid_reduction"))
        metrics["derived_grid_reduction_pct"] = grid_reduction / base_grid if math.isfinite(grid_reduction) and base_grid > 0 else math.nan
        metrics["derived_capacity_reduction"] = _difference(
            baseline.get("public_charger_capacity_kWh_per_slot"),
            metrics.get("public_charger_capacity_kWh_per_slot"),
        )
        metrics["derived_redirection_profit_gain"] = math.nan
        if idx in pair_map:
            no_metrics = pair_map[idx]["no_run"]["metrics"]
            metrics["derived_redirection_profit_gain"] = _difference(metrics.get("annual_profit_SEK"), no_metrics.get("annual_profit_SEK"))
    return pair_map


def _default_label(index: int, run: dict[str, Any]) -> str:
    redir = "With redir." if run["redirection"] == "With redirection" else "No redir."
    return f"R{index + 1:02d}: {run['technology']} | {redir} | {run['vipv_scenario']}"


def _resolve_runs(project_root: Path, values: list[str]) -> list[Path]:
    paths = []
    for value in values:
        path = Path(value)
        paths.append(path.resolve() if path.is_absolute() else (project_root / path).resolve())
    return paths


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Compare two or more completed optimization runs produced by the same solution method. "
            "The workbook emphasizes cross-scenario infrastructure, energy, flexibility, economics, and solver performance."
        )
    )
    parser.add_argument("--method", choices=["auto", "monolithic", "lbbd"], default="auto")
    parser.add_argument("--run", action="append", default=[], help="Completed run folder. Repeat for each run.")
    parser.add_argument("--runs", nargs="+", default=[], help="Alternative: provide multiple run folders after one --runs flag.")
    parser.add_argument("--label", action="append", default=[], help="Optional display label; repeat in the same order as runs.")
    parser.add_argument("--baseline-index", type=int, default=1, help="1-based index of the baseline run (default: first run).")
    parser.add_argument("--out", default=None, help="Optional XLSX output path.")
    parser.add_argument("--project-root", default=str(PROJECT_ROOT))
    parser.add_argument(
        "--skip-input-recompute",
        action="store_true",
        help="Do not reopen demand/parking inputs to reconstruct legacy metrics such as exogenous home-charger count.",
    )
    return parser.parse_args()


def _load_run(run_dir: Path) -> dict[str, Any]:
    if not run_dir.exists():
        raise FileNotFoundError(f"Run folder not found: {run_dir}")
    summary = _load_summary(run_dir)
    metadata = _load_json(run_dir / "run_metadata.json")
    transcript = _read_text(run_dir / "README_RUN.txt")
    method = _detect_method(run_dir, metadata, transcript)
    dataset = _detect_dataset(run_dir, summary, metadata, transcript)
    vipv = _detect_vipv(run_dir, summary, metadata, transcript)
    redirection = _detect_redirection(run_dir, metadata, transcript)
    technology, pv_enabled, bess_enabled = _detect_technology(run_dir, metadata, transcript)
    solver = _solver_metrics(run_dir, method, metadata, transcript)
    return {
        "path": run_dir,
        "metrics": summary,
        "metadata": metadata,
        "method": method,
        "dataset": dataset,
        "vipv_scenario": vipv,
        "redirection": redirection,
        "technology": technology,
        "pv_enabled": pv_enabled,
        "bess_enabled": bess_enabled,
        "solver": solver,
    }


def _validate_methods(runs: list[dict[str, Any]], requested: str) -> str:
    detected = {run["method"] for run in runs}
    if "Unknown" in detected:
        raise ValueError("At least one run method could not be detected from run_metadata.json, README_RUN.txt, or folder naming.")
    if len(detected) != 1:
        raise ValueError(f"This utility requires similar-method runs. Detected methods: {sorted(detected)}")
    method = next(iter(detected))
    if requested != "auto" and method.lower() != requested.lower():
        raise ValueError(f"--method {requested} was requested, but the supplied runs are detected as {method}.")
    return method


def _scenario_summary_rows(runs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for label, key, fmt, rule in SUMMARY_SPEC:
        if rule == "section":
            rows.append({"section": label, "metric": None, "format": "section", "values": [None] * len(runs)})
            continue
        values = []
        for run in runs:
            if not _is_applicable(run, rule):
                values.append(None)
            else:
                value = _finite(run["metrics"].get(key))
                values.append(value if math.isfinite(value) else None)
        rows.append({"section": None, "metric": label, "key": key, "format": fmt, "values": values})
    return rows


SUMMARY_UNITS = {
    "home_chargers_exogenous": "count",
    "chargers_slow_installed": "count",
    "chargers_medium_installed": "count",
    "chargers_fast_installed": "count",
    "public_charger_capacity_kWh_per_slot": "kWh/30 min",
    "PV_panels_installed": "panels",
    "battery_units_installed": "units",
    "grid_total_kWh": "kWh/yr",
    "grid_direct_kWh": "kWh/yr",
    "grid_to_battery_kWh": "kWh/yr",
    "pv_used_total_kWh": "kWh/yr",
    "pv_direct_kWh": "kWh/yr",
    "pv_to_battery_kWh": "kWh/yr",
    "battery_discharge_kWh": "kWh/yr",
    "energy_redirected_kWh": "kWh/yr",
    "redirection_trip_equivalents_annual": "trips/yr",
    "redirection_full_trip_bundles_annual": "bundles/yr",
    "share_redirected_public_demand": "%",
    "derived_pv_use_share": "%",
    "annual_public_demand_kWh": "kWh/yr",
    "annual_home_residual_demand_kWh": "kWh/yr",
    "annual_optimizer_boundary_demand_kWh": "kWh/yr",
    "annual_total_MATSim_charging_demand_kWh": "kWh/yr",
    "annual_slack_kWh": "kWh/yr",
    "derived_slack_share": "%",
    "revenue_all_chargers_SEK": "SEK/yr",
    "grid_cost_SEK": "SEK/yr",
    "redirection_total_cost_SEK": "SEK/yr",
    "redirection_distance_cost_SEK": "SEK/yr",
    "redirection_price_compensation_SEK": "SEK/yr",
    "slack_penalty_SEK": "SEK/yr",
    "capex_chargers_SEK": "SEK/yr",
    "capex_PV_SEK": "SEK/yr",
    "capex_BESS_SEK": "SEK/yr",
    "capex_PV_BESS_SEK": "SEK/yr",
    "annual_profit_SEK": "SEK/yr",
    "derived_profit_gain": "SEK/yr",
    "derived_profit_gain_pct": "%",
    "derived_grid_reduction": "kWh/yr",
    "derived_grid_reduction_pct": "%",
    "derived_capacity_reduction": "kWh/30 min",
    "derived_redirection_profit_gain": "SEK/yr",
}


def _write_summary_sheet(writer: pd.ExcelWriter, runs: list[dict[str, Any]], labels: list[str], baseline_idx: int) -> None:
    wb = writer.book
    ws = wb.create_sheet("Scenario summary")
    n = len(runs)
    last_col = 2 + n
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.cell(1, 1, "Summary of annual infrastructure, energy, flexibility, and economic outcomes across optimization scenarios")
    ws.cell(2, 1, "Method")
    ws.cell(2, 2, runs[0]["method"])
    ws.cell(3, 1, "Selected baseline")
    ws.cell(3, 2, labels[baseline_idx])

    ws.cell(5, 1, "Metric")
    ws.cell(5, 2, "Unit")
    for idx, run in enumerate(runs):
        col = 3 + idx
        ws.cell(5, col, run["technology"])
        ws.cell(6, col, "With redir." if run["redirection"] == "With redirection" else "No redir.")
        ws.cell(7, col, run["vipv_scenario"])
        ws.cell(8, col, labels[idx])
    ws.cell(6, 1, "Redirection")
    ws.cell(7, 1, "VIPV input")
    ws.cell(8, 1, "Run label")

    # Merge contiguous technology headings so a six-run technology/redirection
    # comparison reproduces the two-level structure used in manuscript tables.
    start = 0
    while start < n:
        technology = runs[start]["technology"]
        end = start
        while end + 1 < n and runs[end + 1]["technology"] == technology:
            end += 1
        if end > start:
            ws.merge_cells(start_row=5, start_column=3 + start, end_row=5, end_column=3 + end)
        start = end + 1

    row_idx = 9
    row_formats: dict[int, str] = {}
    for row in _scenario_summary_rows(runs):
        if row["format"] == "section":
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=last_col)
            ws.cell(row_idx, 1, row["section"])
            row_formats[row_idx] = "section"
            row_idx += 1
            continue
        ws.cell(row_idx, 1, row["metric"])
        ws.cell(row_idx, 2, SUMMARY_UNITS.get(row.get("key"), ""))
        for idx, value in enumerate(row["values"]):
            ws.cell(row_idx, 3 + idx, value)
        row_formats[row_idx] = row["format"]
        row_idx += 1

    thin = Side(style="thin", color="B7B7B7")
    dark_fill = PatternFill("solid", fgColor="1F4E78")
    group_fill = PatternFill("solid", fgColor="D9EAF7")
    section_fill = PatternFill("solid", fgColor="D9E1F2")
    baseline_fill = PatternFill("solid", fgColor="FFF2CC")

    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = dark_fill
    ws["A1"].alignment = Alignment(horizontal="center")
    for row in (5, 6, 7, 8):
        for col in range(1, last_col + 1):
            cell = ws.cell(row, col)
            cell.font = Font(bold=True)
            cell.fill = group_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)
    for col in range(3, last_col + 1):
        if col == 3 + baseline_idx:
            for row in range(5, row_idx):
                ws.cell(row, col).fill = baseline_fill if row_formats.get(row) != "section" else section_fill

    for row, fmt in row_formats.items():
        if fmt == "section":
            cell = ws.cell(row, 1)
            cell.font = Font(bold=True)
            cell.fill = section_fill
            continue
        ws.cell(row, 1).alignment = Alignment(horizontal="left")
        ws.cell(row, 2).alignment = Alignment(horizontal="center")
        for col in range(3, last_col + 1):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(horizontal="right")
            if fmt == "count":
                cell.number_format = '#,##0;[Red]-#,##0;0'
            elif fmt == "number":
                cell.number_format = '#,##0;[Red]-#,##0;0'
            elif fmt == "number4":
                cell.number_format = '#,##0.0000;[Red]-#,##0.0000;0.0000'
            elif fmt == "currency":
                cell.number_format = '#,##0;[Red]-#,##0;0'
            elif fmt == "percent":
                cell.number_format = '0.00%;[Red]-0.00%;0.00%'
            elif fmt == "percent6":
                cell.number_format = '0.000000%;[Red]-0.000000%;0.000000%'

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 20
    for col in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 24
    ws.freeze_panes = "C9"
    ws.auto_filter.ref = f"A8:{get_column_letter(last_col)}{row_idx - 1}"


def _baseline_changes_frame(runs: list[dict[str, Any]], labels: list[str], baseline_idx: int) -> pd.DataFrame:
    base = runs[baseline_idx]["metrics"]
    keys = [
        ("Net profit", "annual_profit_SEK"),
        ("Grid electricity purchased", "grid_total_kWh"),
        ("Public charger capacity", "public_charger_capacity_kWh_per_slot"),
        ("Slow chargers", "chargers_slow_installed"),
        ("Medium chargers", "chargers_medium_installed"),
        ("Fast chargers", "chargers_fast_installed"),
        ("PV panels", "PV_panels_installed"),
        ("Battery units", "battery_units_installed"),
        ("PV electricity used locally", "pv_used_total_kWh"),
        ("Battery discharge", "battery_discharge_kWh"),
        ("Annual slack", "annual_slack_kWh"),
    ]
    rows = []
    for idx, run in enumerate(runs):
        for metric_label, key in keys:
            value = _finite(run["metrics"].get(key))
            baseline = _finite(base.get(key))
            rows.append({
                "Run": labels[idx],
                "Metric": metric_label,
                "Value": value if math.isfinite(value) else math.nan,
                "Baseline": baseline if math.isfinite(baseline) else math.nan,
                "Absolute change (run - baseline)": _difference(value, baseline),
                "Relative change": _relative_diff(value, baseline),
            })
    return pd.DataFrame(rows)


def _redirection_effects_frame(runs: list[dict[str, Any]], labels: list[str], pair_map: dict[int, dict[str, Any]]) -> pd.DataFrame:
    metrics = [
        ("Net profit", "annual_profit_SEK"),
        ("Grid electricity purchased", "grid_total_kWh"),
        ("Public charger capacity", "public_charger_capacity_kWh_per_slot"),
        ("Slow chargers", "chargers_slow_installed"),
        ("Medium chargers", "chargers_medium_installed"),
        ("Fast chargers", "chargers_fast_installed"),
        ("PV panels", "PV_panels_installed"),
        ("Battery units", "battery_units_installed"),
        ("PV electricity used locally", "pv_used_total_kWh"),
        ("Battery discharge", "battery_discharge_kWh"),
        ("Redirected energy", "energy_redirected_kWh"),
        ("Redirection incentive cost", "redirection_total_cost_SEK"),
        ("Annual slack", "annual_slack_kWh"),
    ]
    rows = []
    for with_idx, pair in sorted(pair_map.items()):
        no_idx = pair["no_idx"]
        no_metrics = runs[no_idx]["metrics"]
        with_metrics = runs[with_idx]["metrics"]
        for metric_label, key in metrics:
            no_value = _finite(no_metrics.get(key))
            with_value = _finite(with_metrics.get(key))
            rows.append({
                "Dataset": runs[with_idx]["dataset"],
                "VIPV scenario": runs[with_idx]["vipv_scenario"],
                "Technology": runs[with_idx]["technology"],
                "No-redirection run": labels[no_idx],
                "With-redirection run": labels[with_idx],
                "Metric": metric_label,
                "No redirection": no_value if math.isfinite(no_value) else math.nan,
                "With redirection": with_value if math.isfinite(with_value) else math.nan,
                "Absolute effect (with - no)": _difference(with_value, no_value),
                "Relative effect": _relative_diff(with_value, no_value),
            })
    return pd.DataFrame(rows)


def _metadata_frame(runs: list[dict[str, Any]], labels: list[str]) -> pd.DataFrame:
    rows = []
    for idx, run in enumerate(runs):
        solver = run["solver"]
        rows.append({
            "Run": labels[idx],
            "Method": run["method"],
            "Dataset": run["dataset"],
            "VIPV scenario": run["vipv_scenario"],
            "Technology": run["technology"],
            "PV enabled": run["pv_enabled"],
            "BESS enabled": run["bess_enabled"],
            "Redirection": run["redirection"],
            "Requested convergence gap (%)": solver.get("requested_gap_pct"),
            "Achieved/certified gap (%)": solver.get("achieved_gap_pct"),
            "Termination": solver.get("termination"),
            "Solve/decomposition time (s)": solver.get("solve_seconds"),
            "Total runtime (s)": solver.get("total_runtime_s"),
            "Peak process-tree RSS (MB)": solver.get("peak_rss_MB"),
            "Threads": solver.get("threads"),
            "Run folder": str(run["path"]),
        })
    return pd.DataFrame(rows)


def _raw_metrics_frame(runs: list[dict[str, Any]], labels: list[str]) -> pd.DataFrame:
    keys = sorted(set().union(*(set(run["metrics"].keys()) for run in runs)))
    rows = []
    for key in keys:
        row = {"Metric": key}
        for idx, run in enumerate(runs):
            row[labels[idx]] = run["metrics"].get(key, math.nan)
        rows.append(row)
    return pd.DataFrame(rows)


def _ranking_frame(runs: list[dict[str, Any]], labels: list[str]) -> pd.DataFrame:
    frame = pd.DataFrame({
        "Run": labels,
        "Net profit (SEK/yr)": [_finite(r["metrics"].get("annual_profit_SEK")) for r in runs],
        "Grid purchase (kWh/yr)": [_finite(r["metrics"].get("grid_total_kWh")) for r in runs],
        "Public charger capacity (kWh/30 min)": [_finite(r["metrics"].get("public_charger_capacity_kWh_per_slot")) for r in runs],
        "PV use share": [_finite(r["metrics"].get("derived_pv_use_share")) for r in runs],
        "Annual slack (kWh/yr)": [_finite(r["metrics"].get("annual_slack_kWh")) for r in runs],
    })
    frame["Profit rank (1=highest)"] = frame["Net profit (SEK/yr)"].rank(ascending=False, method="min")
    frame["Grid rank (1=lowest)"] = frame["Grid purchase (kWh/yr)"].rank(ascending=True, method="min")
    frame["Capacity rank (1=lowest)"] = frame["Public charger capacity (kWh/30 min)"].rank(ascending=True, method="min")
    frame["PV-use rank (1=highest)"] = frame["PV use share"].rank(ascending=False, method="min")
    frame["Slack rank (1=lowest)"] = frame["Annual slack (kWh/yr)"].rank(ascending=True, method="min")
    return frame


def _consistency_frame(runs: list[dict[str, Any]], method: str) -> pd.DataFrame:
    datasets = sorted({r["dataset"] for r in runs})
    gaps = [_finite(r["solver"].get("requested_gap_pct")) for r in runs]
    finite_gaps = [g for g in gaps if math.isfinite(g)]
    checks = [
        {
            "Check": "Same solution method",
            "Status": "PASS",
            "Details": method,
        },
        {
            "Check": "Same dataset size",
            "Status": "PASS" if len(datasets) == 1 else "WARNING",
            "Details": ", ".join(datasets),
        },
        {
            "Check": "Requested convergence gaps comparable",
            "Status": "PASS" if len({round(g, 8) for g in finite_gaps}) <= 1 else "WARNING",
            "Details": ", ".join(f"{g:.6g}%" for g in finite_gaps) if finite_gaps else "Not available",
        },
        {
            "Check": "Run count",
            "Status": "PASS",
            "Details": str(len(runs)),
        },
    ]
    vipv = sorted({r["vipv_scenario"] for r in runs})
    checks.append({
        "Check": "VIPV input scenarios",
        "Status": "INFO",
        "Details": ", ".join(vipv),
    })
    return pd.DataFrame(checks)


def _style_standard_sheet(ws, percent_columns: set[int] | None = None) -> None:
    percent_columns = percent_columns or set()
    dark_fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="D9D9D9")
    for cell in ws[1]:
        cell.fill = dark_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        max_len = 0
        for cell in col[: min(ws.max_row, 200)]:
            text = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(text))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 42)
        if col_idx in percent_columns:
            for cell in col[1:]:
                cell.number_format = "0.00%"


def _style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    for sheet_name in wb.sheetnames:
        if sheet_name == "Scenario summary":
            continue
        ws = wb[sheet_name]
        pct_cols = set()
        for cell in ws[1]:
            if cell.value and ("Relative" in str(cell.value) or "share" in str(cell.value).lower()):
                pct_cols.add(cell.column)
        _style_standard_sheet(ws, pct_cols)
    wb.save(path)


def main() -> int:
    args = parse_args()
    project_root = Path(args.project_root).resolve()
    run_values = list(args.run) + list(args.runs)
    if len(run_values) < 2:
        raise SystemExit("At least two completed run folders are required. Use repeated --run flags or --runs path1 path2 ...")
    run_paths = _resolve_runs(project_root, run_values)
    runs = [_load_run(path) for path in run_paths]
    method = _validate_methods(runs, args.method)

    if not (1 <= args.baseline_index <= len(runs)):
        raise SystemExit(f"--baseline-index must be between 1 and {len(runs)}")
    baseline_idx = args.baseline_index - 1

    if args.label:
        if len(args.label) != len(runs):
            raise SystemExit("If --label is used, supply exactly one --label for every run folder.")
        labels = list(args.label)
    else:
        labels = [_default_label(i, run) for i, run in enumerate(runs)]

    if len(set(labels)) != len(labels):
        raise SystemExit("Scenario labels must be unique.")

    home_cache: dict[tuple[str, str], float] = {}
    for run in runs:
        _infer_missing_metrics(project_root, run, home_cache, args.skip_input_recompute)
    pair_map = _add_derived_comparisons(runs, baseline_idx)

    if args.out:
        output = Path(args.out)
        output = output.resolve() if output.is_absolute() else (project_root / output).resolve()
    else:
        out_dir = project_root / "runs" / "comparisons"
        out_dir.mkdir(parents=True, exist_ok=True)
        output = out_dir / f"{datetime.now():%Y-%m-%d_%H%M%S}_{method.lower()}_scenario_comparison.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)

    baseline_changes = _baseline_changes_frame(runs, labels, baseline_idx)
    redirection_effects = _redirection_effects_frame(runs, labels, pair_map)
    metadata = _metadata_frame(runs, labels)
    rankings = _ranking_frame(runs, labels)
    raw = _raw_metrics_frame(runs, labels)
    consistency = _consistency_frame(runs, method)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _write_summary_sheet(writer, runs, labels, baseline_idx)
        baseline_changes.to_excel(writer, sheet_name="Baseline changes", index=False)
        if redirection_effects.empty:
            pd.DataFrame([{
                "Information": "No unique no-redirection/with-redirection run pairs were found for identical dataset, VIPV input, and technology."
            }]).to_excel(writer, sheet_name="Redirection effects", index=False)
        else:
            redirection_effects.to_excel(writer, sheet_name="Redirection effects", index=False)
        rankings.to_excel(writer, sheet_name="Scenario ranking", index=False)
        metadata.to_excel(writer, sheet_name="Run metadata", index=False)
        raw.to_excel(writer, sheet_name="Raw metrics", index=False)
        consistency.to_excel(writer, sheet_name="Consistency checks", index=False)

    _style_workbook(output)

    print(f"Same-method scenario comparison complete ({method}).")
    print(f"Runs compared: {len(runs)}")
    print(f"Baseline: {labels[baseline_idx]}")
    print(f"Workbook written to: {output}")
    if pair_map:
        print(f"Matched no-/with-redirection pairs: {len(pair_map)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
