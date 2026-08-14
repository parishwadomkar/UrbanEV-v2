#!/usr/bin/env python
from __future__ import annotations

import argparse
import math
import json
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from computational_complexity import (
    collect_solver_log_detail,
    preprocessing_scalars,
    read_scalar_file,
    slot_redirection_complexity,
    summarize_solver_logs,
)
from data_loader import load_inputs
from preprocessing import preprocess
from input_scenarios import VIPV_SCENARIOS, resolve_scenario_paths

PROJECT_ROOT = Path(__file__).resolve().parents[1]
MONTH_DAYS = {
    "January": 31, "February": 28, "March": 31, "April": 30,
    "May": 31, "June": 30, "July": 31, "August": 31,
    "September": 30, "October": 31, "November": 30, "December": 31,
}


def load_json(path: Path) -> dict:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def resolve_project_path(project_root: Path, value: str | Path) -> Path:
    path = Path(value)
    return path.resolve() if path.is_absolute() else (project_root / path).resolve()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compare monolithic and LBBD result folders."
    )
    parser.add_argument("--monolithic-run", required=True, help="Monolithic run folder.")
    parser.add_argument("--lbbd-run", required=True, help="LBBD run folder.")
    parser.add_argument("--out", default=None, help="Optional output XLSX path.")
    parser.add_argument("--project-root", default=str(PROJECT_ROOT))
    parser.add_argument("--dataset", choices=["small", "full"], default=None)
    parser.add_argument("--scenario", choices=["with_redirection", "no_redirection"], default=None)
    parser.add_argument("--vipv-scenario", choices=VIPV_SCENARIOS, default=None)
    parser.add_argument(
        "--skip-complexity-recompute", action="store_true",
        help="Do not reload project inputs to recompute exact preprocessing/redirection set sizes.",
    )
    return parser.parse_args()


def _resolve(root: Path, value: str) -> Path:
    path = Path(value)
    return path.resolve() if path.is_absolute() else (root / path).resolve()


def _number(value: Any) -> Any:
    if value is None:
        return math.nan
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return math.nan
    try:
        return float(text)
    except ValueError:
        return text


def _read_summary(run_dir: Path) -> dict[str, Any]:
    path = run_dir / "results" / "model_summary.csv"
    if not path.exists():
        raise FileNotFoundError(f"Missing model summary: {path}")
    frame = pd.read_csv(path)
    if not {"Metric", "Value"}.issubset(frame.columns):
        raise ValueError(f"Expected Metric/Value columns in {path}")
    return {str(row.Metric): _number(row.Value) for row in frame.itertuples(index=False)}


def _read_csv_if_present(path: Path) -> pd.DataFrame | None:
    if not path.exists() or path.stat().st_size == 0:
        return None
    try:
        return pd.read_csv(path)
    except pd.errors.EmptyDataError:
        return None


def _annual_public_demand(run_dir: Path) -> float:
    frame = _read_csv_if_present(run_dir / "results" / "hourly_energy.csv")
    if frame is None or "Demand_public_base_kWh_day" not in frame.columns:
        return math.nan
    if "Month" not in frame.columns:
        return math.nan
    days = frame["Month"].map(MONTH_DAYS).fillna(0)
    return float((pd.to_numeric(frame["Demand_public_base_kWh_day"], errors="coerce").fillna(0) * days).sum())


def _annual_slack(run_dir: Path) -> float:
    frame = _read_csv_if_present(run_dir / "results" / "slack.csv")
    if frame is None or frame.empty:
        return 0.0
    for column in ("Slack_kWh_annual", "AnnualSlack_kWh", "annual_slack_kWh"):
        if column in frame.columns:
            return float(pd.to_numeric(frame[column], errors="coerce").fillna(0).sum())
    return math.nan


def _active_redirection_arcs(run_dir: Path) -> float:
    frame = _read_csv_if_present(run_dir / "results" / "redirections.csv")
    if frame is None or frame.empty:
        return 0.0
    if "Energy_kWh_day" in frame.columns:
        return float((pd.to_numeric(frame["Energy_kWh_day"], errors="coerce").fillna(0).abs() > 1e-8).sum())
    return float(len(frame))


def _load_capacity_config(project_root: Path) -> dict[str, float]:
    import json

    path = project_root / "config" / "model_config.json"
    cfg = json.loads(path.read_text(encoding="utf-8"))
    return {str(k): float(v) for k, v in cfg["charger_capacity_kwh_per_slot"].items()}


def _enrich(metrics: dict[str, Any], run_dir: Path, capacity: dict[str, float]) -> dict[str, Any]:
    out = dict(metrics)
    if not math.isfinite(float(out.get("annual_public_demand_kWh", math.nan))):
        out["annual_public_demand_kWh"] = _annual_public_demand(run_dir)
    if not math.isfinite(float(out.get("annual_slack_kWh", math.nan))):
        out["annual_slack_kWh"] = _annual_slack(run_dir)
    if not math.isfinite(float(out.get("active_redirection_arcs", math.nan))):
        out["active_redirection_arcs"] = _active_redirection_arcs(run_dir)
    if not math.isfinite(float(out.get("public_charger_capacity_kWh_per_slot", math.nan))):
        out["public_charger_capacity_kWh_per_slot"] = sum(
            float(out.get(f"chargers_{kind}_installed", 0.0) or 0.0) * float(capacity[kind])
            for kind in ("slow", "medium", "fast")
        )
    demand = float(out.get("annual_public_demand_kWh", math.nan))
    redirected = float(out.get("energy_redirected_kWh", math.nan))
    if not math.isfinite(float(out.get("share_redirected_public_demand", math.nan))):
        out["share_redirected_public_demand"] = redirected / demand if demand > 0 else 0.0
    return out



def _read_json_if_present(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _infer_dataset(run_dirs: dict[str, Path], explicit: str | None) -> str:
    if explicit:
        return explicit
    for run_dir in run_dirs.values():
        candidates = [
            run_dir / "run_metadata.json",
            run_dir / "logs" / "lbbd_manifest.json",
        ]
        for path in candidates:
            value = _read_json_if_present(path).get("dataset")
            if value in {"small", "full"}:
                return str(value)
        lower = run_dir.name.lower()
        for value in ("small", "full"):
            if f"_{value}_" in lower:
                return value
    raise ValueError("Could not infer a common dataset; provide --dataset small or --dataset full.")


def _infer_scenario(run_dirs: dict[str, Path], explicit: str | None) -> str:
    if explicit:
        return explicit
    for run_dir in run_dirs.values():
        for path in [run_dir / "run_metadata.json", run_dir / "logs" / "lbbd_manifest.json"]:
            value = _read_json_if_present(path).get("scenario")
            if value in {"with_redirection", "no_redirection"}:
                return str(value)
        lower = run_dir.name.lower()
        if "no_redirection" in lower or "noredirection" in lower:
            return "no_redirection"
        if "with_redirection" in lower or "withredirection" in lower:
            return "with_redirection"
    return "with_redirection"



def _infer_vipv_scenario(run_dirs: dict[str, Path], explicit: str | None) -> str:
    if explicit:
        return explicit
    for run_dir in run_dirs.values():
        for path in [run_dir / "run_metadata.json", run_dir / "logs" / "lbbd_manifest.json"]:
            value = _read_json_if_present(path).get("vipv_scenario")
            if value in VIPV_SCENARIOS:
                return str(value)
        summary = _read_summary(run_dir)
        value = summary.get("vipv_scenario")
        if value in VIPV_SCENARIOS:
            return str(value)
        for candidate in VIPV_SCENARIOS:
            if f"_{candidate.lower()}_" in run_dir.name.lower():
                return candidate
    return "noVIPV"

def _load_preprocessed_comparison_data(root: Path, dataset: str, scenario: str, vipv_scenario: str) -> dict:
    model_cfg = load_json(root / "config" / "model_config.json")
    paths = resolve_scenario_paths(root, dataset, vipv_scenario)
    raw = load_inputs(paths)
    data = preprocess(raw, model_cfg)
    data["dataset"] = dataset
    data["vipv_scenario"] = vipv_scenario
    if scenario == "no_redirection":
        data["allowed"] = []
        data["allowed_st"] = []
        data["OUT"] = {}
        data["IN"] = {}
        data["ORIGIN_ST"] = []
        data["DEST_ST"] = []
    return data



def _history_total_seconds(run_dir: Path) -> float:
    for path in [
        run_dir / "results" / "lbbd_history.csv",
    ]:
        frame = _read_csv_if_present(path)
        if frame is not None and "elapsed_seconds" in frame.columns:
            values = pd.to_numeric(frame["elapsed_seconds"], errors="coerce").dropna()
            if not values.empty:
                return float(values.max())
    meta = _read_json_if_present(run_dir / "run_metadata.json")
    for key in ("elapsed_seconds", "total_runtime_seconds"):
        try:
            value = float(meta.get(key))
            if math.isfinite(value):
                return value
        except Exception:
            pass
    return math.nan


def _complexity_frames(run_dirs: dict[str, Path]) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    all_detail = []
    all_summary = []
    scalar_by_method: dict[str, dict[str, Any]] = {}
    for method, run_dir in run_dirs.items():
        scalars = read_scalar_file(run_dir)
        if not math.isfinite(float(scalars.get("total_runtime_seconds", math.nan))):
            scalars["total_runtime_seconds"] = _history_total_seconds(run_dir)
        scalar_by_method[method] = scalars
        detail = collect_solver_log_detail(run_dir)
        if not detail.empty:
            detail.insert(0, "Method", method)
            all_detail.append(detail)
            summary = summarize_solver_logs(detail.drop(columns=["Method"]).copy())
            if not summary.empty:
                summary.insert(0, "Method", method)
                all_summary.append(summary)

    detail_frame = pd.concat(all_detail, ignore_index=True) if all_detail else pd.DataFrame()
    summary_frame = pd.concat(all_summary, ignore_index=True) if all_summary else pd.DataFrame()
    timing_keys = sorted(set().union(*[
        {key for key in metrics if key.endswith("_seconds")}
        for metrics in scalar_by_method.values()
    ]))
    timing_rows = []
    for key in timing_keys:
        row = {"Phase": key}
        for method in ("Monolithic", "LBBD"):
            row[method] = scalar_by_method.get(method, {}).get(key, math.nan)
        timing_rows.append(row)
    timing_frame = pd.DataFrame(timing_rows)

    primary_role = {"Monolithic": "main_model", "LBBD": "master"}
    summary_metrics = [
        ("Total runtime", "total_runtime_seconds", "s"),
        ("Peak process-tree RSS", "peak_process_tree_rss_MB", "MB"),
        ("Input loading", "input_load_seconds", "s"),
        ("Preprocessing", "preprocessing_seconds", "s"),
        ("Pyomo/master build", None, "s"),
        ("Initial solver rows", "Max initial_rows", "count"),
        ("Initial solver columns", "Max initial_columns", "count"),
        ("Initial solver nonzeros", "Max initial_nonzeros", "count"),
        ("Presolved rows", "Max presolved_rows", "count"),
        ("Presolved columns", "Max presolved_columns", "count"),
        ("Presolved nonzeros", "Max presolved_nonzeros", "count"),
        ("Rows removed by presolve", "Rows removed by presolve", "fraction"),
        ("Columns removed by presolve", "Columns removed by presolve", "fraction"),
        ("Solver calls for primary role", "Solver calls", "count"),
        ("Total primary-role solver time", "Total solver_seconds", "s"),
        ("Total primary-role nodes", "Total nodes_explored", "count"),
        ("Total primary-role work units", "Total work_units", "work units"),
        ("Maximum reported barrier factor memory", "Max barrier_factor_memory_MB", "MB"),
    ]
    computational_rows = []
    for label, key, unit in summary_metrics:
        row = {"Metric": label, "Unit": unit}
        for method in ("Monolithic", "LBBD"):
            scalars = scalar_by_method.get(method, {})
            if label == "Pyomo/master build":
                if method == "LBBD":
                    value = scalars.get("master_build_seconds", math.nan)
                else:
                    value = scalars.get("model_build_seconds", math.nan)
            elif key in scalars:
                value = scalars.get(key, math.nan)
            else:
                value = math.nan
                if not summary_frame.empty:
                    subset = summary_frame[
                        (summary_frame["Method"] == method)
                        & (summary_frame["Model role"] == primary_role[method])
                    ]
                    if not subset.empty and key in subset.columns:
                        value = subset.iloc[0][key]
            row[method] = value
        computational_rows.append(row)
    computational_frame = pd.DataFrame(computational_rows)
    return computational_frame, timing_frame, summary_frame, detail_frame


ROW_SPEC = [
    ("Slow chargers (11 kW)", "chargers_slow_installed", 0, "count"),
    ("Medium chargers (22 kW)", "chargers_medium_installed", 0, "count"),
    ("Fast chargers (50 kW)", "chargers_fast_installed", 0, "count"),
    ("Public charger capacity (kWh/30 min)", "public_charger_capacity_kWh_per_slot", 0, "number"),
    ("PV panels (500 W)", "PV_panels_installed", 0, "count"),
    ("Battery units (10 kWh)", "battery_units_installed", 0, "count"),
    ("Energy Flows (kWh/yr)", None, 0, "section"),
    ("Grid electricity purchase", "grid_total_kWh", 0, "number"),
    ("Direct grid to chargers", "grid_direct_kWh", 1, "number"),
    ("Grid to battery charging", "grid_to_battery_kWh", 1, "number"),
    ("PV generated locally", "pv_used_total_kWh", 0, "number"),
    ("Direct to chargers", "pv_direct_kWh", 1, "number"),
    ("PV to battery charging", "pv_to_battery_kWh", 1, "number"),
    ("Battery Discharge", "battery_discharge_kWh", 0, "number"),
    ("Public demand redirected", "energy_redirected_kWh", 0, "number"),
    ("Active redirection arcs", "active_redirection_arcs", 0, "count"),
    ("Share of redirected public demand", "share_redirected_public_demand", 0, "percent"),
    ("Annual economic terms (SEK/yr)", None, 0, "section"),
    ("Public charging revenue", "revenue_all_chargers_SEK", 0, "currency"),
    ("Grid electricity cost", "grid_cost_SEK", 0, "currency"),
    ("Redirection incentive", "redirection_total_cost_SEK", 0, "currency"),
    ("Redirection distance compensation", "redirection_distance_cost_SEK", 1, "currency"),
    ("Redirection price compensation", "redirection_price_compensation_SEK", 1, "currency"),
    ("Unmet demand penalty", "slack_penalty_SEK", 0, "currency"),
    ("Annual CapEx chargers", "capex_chargers_SEK", 0, "currency"),
    ("Annual CapEx (PV + BESS)", "capex_PV_BESS_SEK", 0, "currency"),
    ("Net Profit (Obj)", "annual_profit_SEK", 0, "currency"),
]


def _difference(value: Any, base: Any) -> float:
    try:
        a, b = float(value), float(base)
        return a - b if math.isfinite(a) and math.isfinite(b) else math.nan
    except Exception:
        return math.nan


def _relative(value: Any, base: Any) -> float:
    diff = _difference(value, base)
    try:
        b = float(base)
        if math.isfinite(diff) and math.isfinite(b) and abs(b) > 1e-12:
            return diff / abs(b)
    except Exception:
        pass
    return math.nan


def _requested_frame(all_metrics: dict[str, dict[str, Any]]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    mono = all_metrics["Monolithic"]
    for label, key, indent, kind in ROW_SPEC:
        record: dict[str, Any] = {
            "Metric": label,
            "MetricKey": key or "",
            "Indent": indent,
            "Format": kind,
        }
        if key is not None:
            for method in ("Monolithic", "LBBD"):
                record[method] = all_metrics[method].get(key, math.nan)
            record["LBBD - Monolithic"] = _difference(record["LBBD"], record["Monolithic"])
            record["LBBD relative difference"] = _relative(record["LBBD"], record["Monolithic"])
        rows.append(record)
    return pd.DataFrame(rows)


def _raw_frame(all_metrics: dict[str, dict[str, Any]]) -> pd.DataFrame:
    keys = sorted(set().union(*(set(values) for values in all_metrics.values())))
    rows = []
    for key in keys:
        row = {"Metric": key}
        for method in ("Monolithic", "LBBD"):
            row[method] = all_metrics[method].get(key, math.nan)
        row["LBBD - Monolithic"] = _difference(row["LBBD"], row["Monolithic"])
        row["LBBD relative difference"] = _relative(row["LBBD"], row["Monolithic"])
        rows.append(row)
    return pd.DataFrame(rows)



def _finite_float(value: Any) -> float:
    try:
        number = float(value)
        return number if math.isfinite(number) else math.nan
    except Exception:
        return math.nan


def _first_scalar(scalars: dict[str, Any], *keys: str) -> float:
    for key in keys:
        value = _finite_float(scalars.get(key))
        if math.isfinite(value):
            return value
    return math.nan


def _solver_summary_value(
    summary: pd.DataFrame,
    method: str,
    roles: tuple[str, ...],
    column: str,
) -> float:
    if summary is None or summary.empty or column not in summary.columns:
        return math.nan
    subset = summary[
        (summary["Method"].astype(str) == str(method))
        & (summary["Model role"].astype(str).isin(list(roles)))
    ]
    if subset.empty:
        return math.nan
    values = pd.to_numeric(subset[column], errors="coerce").dropna()
    return float(values.sum()) if column.startswith("Total ") or column == "Solver calls" else (
        float(values.max()) if not values.empty else math.nan
    )


def _read_history(run_dir: Path) -> pd.DataFrame:
    candidates = [
        run_dir / "results" / "lbbd_history.csv",
        run_dir / "results" / "alternative_c_history.csv",
        run_dir / "iterations" / "lbbd_iteration_history.csv",
        run_dir / "results" / "lbbd_iteration_history.csv",
    ]
    for path in candidates:
        frame = _read_csv_if_present(path)
        if frame is not None:
            return frame
    return pd.DataFrame()


def _slot_structure_statistics(data: dict) -> dict[str, float]:
    months = [str(v) for v in data.get("MONTHS", [])]
    slots = [int(v) for v in data.get("INTERVALS", [])]
    cells = [int(v) for v in data.get("hex_ids", [])]
    charger_types = len(data.get("PUB_TYPES", []))
    by_slot: dict[tuple[str, int], list[tuple[int, int]]] = {
        (m, t): [] for m in months for t in slots
    }
    for i, j, mon, t in data.get("allowed_st", []):
        by_slot.setdefault((str(mon), int(t)), []).append((int(i), int(j)))

    arc_counts: list[int] = []
    origin_counts: list[int] = []
    destination_counts: list[int] = []
    equality_counts: list[int] = []
    type_pair_counts: list[int] = []
    component_counts: list[int] = []

    for key in [(m, t) for m in months for t in slots]:
        arcs = by_slot.get(key, [])
        origins = {i for i, _ in arcs}
        destinations = {j for _, j in arcs}
        arc_count = len(arcs)
        arc_counts.append(arc_count)
        origin_counts.append(len(origins))
        destination_counts.append(len(destinations))
        type_pair_counts.append(arc_count * charger_types * charger_types)
        equality_counts.append(
            charger_types * (len(origins) + len(destinations)) + arc_count
        )

        parent = {i: i for i in cells}
        def find(a: int) -> int:
            while parent[a] != a:
                parent[a] = parent[parent[a]]
                a = parent[a]
            return a
        def union(a: int, b: int) -> None:
            ra, rb = find(a), find(b)
            if ra != rb:
                parent[max(ra, rb)] = min(ra, rb)
        for i, j in arcs:
            union(i, j)
        component_counts.append(len({find(i) for i in cells}))

    def avg(values: list[int]) -> float:
        return float(sum(values) / len(values)) if values else 0.0
    def maximum(values: list[int]) -> int:
        return int(max(values)) if values else 0

    return {
        "I": len(cells),
        "M": len(months),
        "H": len(slots),
        "HSOC": len(data.get("HSOC", [])),
        "C": charger_types,
        "B": len(data.get("DEMAND_CLASSES", [])),
        "blocks": len(months) * len(slots),
        "A": len(data.get("allowed_st", [])),
        "O": len(data.get("ORIGIN_ST", [])),
        "D": len(data.get("DEST_ST", [])),
        "mean_arcs": avg(arc_counts),
        "max_arcs": maximum(arc_counts),
        "sum_type_pairs": float(sum(type_pair_counts)),
        "mean_type_pairs": avg(type_pair_counts),
        "max_type_pairs": maximum(type_pair_counts),
        "sum_slot_equalities": float(sum(equality_counts)),
        "mean_slot_equalities": avg(equality_counts),
        "max_slot_equalities": maximum(equality_counts),
        "sum_components": float(sum(component_counts)),
        "mean_components": avg(component_counts),
        "max_components": maximum(component_counts),
    }


def _complexity_record(
    component: str,
    metric: str,
    definition: str,
    current_size: Any,
    unit: str = "count",
    notes: str = "",
) -> dict[str, Any]:
    return {
        "Component": component,
        "Metric": metric,
        "Order / definition": definition,
        "Current size": current_size,
        "Unit": unit,
        "Source / notes": notes,
    }


def _detailed_complexity_tables(
    data: dict | None,
    run_dirs: dict[str, Path],
    solver_summary: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    scalar_map = {method: read_scalar_file(path) for method, path in run_dirs.items()}
    stats = _slot_structure_statistics(data) if data is not None else {
        key: math.nan for key in [
            "I", "M", "H", "HSOC", "C", "B", "blocks", "A", "O", "D",
            "mean_arcs", "max_arcs", "sum_type_pairs", "mean_type_pairs",
            "max_type_pairs", "sum_slot_equalities", "mean_slot_equalities",
            "max_slot_equalities", "sum_components", "mean_components", "max_components",
        ]
    }
    I, M, H, HSOC, C, B, A = (stats[k] for k in ["I", "M", "H", "HSOC", "C", "B", "A"])

    lsc = scalar_map.get("LBBD", {})
    lh = _read_history(run_dirs["LBBD"])
    l_cuts = (
        float(pd.to_numeric(lh.get("new_cuts_total"), errors="coerce").fillna(0).sum())
        if not lh.empty and "new_cuts_total" in lh.columns else math.nan
    )
    embedded_energy = (
        4 * I * M * H + I * M * HSOC
        if all(math.isfinite(v) for v in [I, M, H, HSOC])
        else math.nan
    )
    l_rows = [
        _complexity_record("Initial LBBD master", "Binary variables", "Bit encodings for exact configuration cuts", _first_scalar(lsc, "initial_master_variables_binary")),
        _complexity_record("Initial LBBD master", "Non-binary integer variables", "|I|(|C|+2): chargers, PV panels and BESS units", _first_scalar(lsc, "initial_master_variables_integer_nonbinary")),
        _complexity_record("Initial LBBD master", "Continuous variables", "Embedded redirection, service, Theta, slack and linked energy relaxation", _first_scalar(lsc, "initial_master_variables_continuous")),
        _complexity_record("Initial LBBD master", "Total variables", "Recorded active Pyomo variables", _first_scalar(lsc, "initial_master_variables_total")),
        _complexity_record("Initial LBBD master", "Active constraints", "Recorded active Pyomo constraints", _first_scalar(lsc, "initial_master_constraints_active")),
        _complexity_record("Final LBBD master", "Binary variables", "Initial bit encodings plus pooled partial-cut indicators", _first_scalar(lsc, "final_master_variables_binary")),
        _complexity_record("Final LBBD master", "Total variables", "Master after all generated cuts and pooled indicators", _first_scalar(lsc, "final_master_variables_total")),
        _complexity_record("Final LBBD master", "Active constraints", "Master after all generated cuts and links", _first_scalar(lsc, "final_master_constraints_active")),
        _complexity_record("Final LBBD master", "Cuts added", "sum_k new inference cuts", l_cuts),
        _complexity_record("Embedded recourse relaxation", "Continuous aggregate redirection variables", "|A|", A),
        _complexity_record("Embedded recourse relaxation", "Charger-type service variables", "|I||M||H||C|", I * M * H * C if all(math.isfinite(v) for v in [I, M, H, C]) else math.nan),
        _complexity_record("Embedded recourse relaxation", "Home/local/slack variables", "4|I||M||H|", 4 * I * M * H if all(math.isfinite(v) for v in [I, M, H]) else math.nan),
        _complexity_record("Embedded recourse relaxation", "Component-slot Theta variables", "sum_(m,t) number of connected components", stats["sum_components"]),
        _complexity_record("Embedded recourse relaxation", "Linked PV/BESS variables", "4|I||M||H| + |I||M||H_soc|", embedded_energy),
        _complexity_record("Exact logic omitted from master", "Type-pair arc flows", "|A||C|^2", A * C * C),
        _complexity_record("Exact logic omitted from master", "Activation, trip and tail variables", "3|A|", 3 * A),
        _complexity_record("Exact logic omitted from master", "Total exact redirection variables avoided in master", "|A|(|C|^2+3)", A * (C * C + 3)),
        _complexity_record("Exact annual certification MIP", "Binary variables", "Recorded exact fixed-investment annual model", _first_scalar(lsc, "best_exact_annual_oracle_variables_binary")),
        _complexity_record("Exact annual certification MIP", "Non-binary integer variables", "Recorded exact fixed-investment annual model", _first_scalar(lsc, "best_exact_annual_oracle_variables_integer_nonbinary")),
        _complexity_record("Exact annual certification MIP", "Continuous variables", "Recorded exact fixed-investment annual model", _first_scalar(lsc, "best_exact_annual_oracle_variables_continuous")),
        _complexity_record("Exact annual certification MIP", "Active constraints", "Recorded exact fixed-investment annual model", _first_scalar(lsc, "best_exact_annual_oracle_constraints_active")),
        _complexity_record("Oracle system", "Connected component-slot models", "sum_(m,t)|K_(m,t)|", stats["sum_components"]),
        _complexity_record("Oracle system", "Mean / maximum components per slot", "mean / max |K_(m,t)|", f"{stats['mean_components']:.2f} / {int(stats['max_components'])}" if math.isfinite(stats["mean_components"]) else ""),
        _complexity_record("Oracle system", "Exact annual MIP solver calls", "Observed exact certification solves", _solver_summary_value(solver_summary, "LBBD", ("exact_annual_oracle",), "Solver calls")),
        _complexity_record("Oracle system", "Linked annual LP solver calls", "Candidate and core-point LP evaluations", _solver_summary_value(solver_summary, "LBBD", ("linked_annual_lp",), "Solver calls")),
        _complexity_record("Oracle system", "Monthly logic-MIP solver calls", "Exact monthly partial-assignment inference", _solver_summary_value(solver_summary, "LBBD", ("monthly_logic_mip",), "Solver calls")),
        _complexity_record("Oracle system", "Component-LP solver calls", "Optional component multicuts", _solver_summary_value(solver_summary, "LBBD", ("component_lp",), "Solver calls")),
        _complexity_record("Observed execution", "Iterations completed", "Recorded LBBD iterations", _first_scalar(lsc, "iterations_completed")),
        _complexity_record("Observed execution", "Distinct exact investment candidates", "Exact annual oracle cache size", _first_scalar(lsc, "unique_exact_candidates")),
        _complexity_record("Observed execution", "Master build time", "Embedded master construction", _first_scalar(lsc, "master_build_seconds"), unit="s"),
        _complexity_record("Observed execution", "Oracle-model build time", "Monthly/annual reusable model construction", _first_scalar(lsc, "oracle_model_build_seconds"), unit="s"),
        _complexity_record("Observed execution", "Total runtime", "Build + decomposition + export + figures", _first_scalar(lsc, "total_runtime_seconds"), unit="s"),
        _complexity_record("Observed execution", "Peak process-tree RSS", "Python plus descendant Gurobi processes", _first_scalar(lsc, "peak_process_tree_rss_MB"), unit="MB"),
    ]

    def primary_scalar(method: str, suffix: str) -> float:
        prefix = {"Monolithic": "main_model", "LBBD": "initial_master"}[method]
        return _first_scalar(scalar_map.get(method, {}), f"{prefix}_{suffix}")

    def primary_solver(method: str, column: str) -> float:
        roles = {"Monolithic": ("main_model",), "LBBD": ("master",)}[method]
        return _solver_summary_value(solver_summary, method, roles, column)

    def build_seconds(method: str) -> float:
        key = "master_build_seconds" if method == "LBBD" else "model_build_seconds"
        return _first_scalar(scalar_map.get(method, {}), key)

    metric_specs = [
        ("Total runtime", "s", "Lower is better", lambda m: _first_scalar(scalar_map.get(m, {}), "total_runtime_seconds"), "End-to-end wall-clock time."),
        ("Peak process-tree RSS", "MB", "Lower is better", lambda m: _first_scalar(scalar_map.get(m, {}), "peak_process_tree_rss_MB"), "Measured Python plus descendant solver RSS; requires psutil."),
        ("Primary model build time", "s", "Lower is better", build_seconds, "Monolithic model build; LBBD embedded-master build."),
        ("Primary model total variables", "count", "Lower is better", lambda m: primary_scalar(m, "variables_total"), "Initial model presented by each method."),
        ("Primary model binary variables", "count", "Lower is better", lambda m: primary_scalar(m, "variables_binary"), "Initial active Pyomo binary variables."),
        ("Primary model general-integer variables", "count", "Lower is better", lambda m: primary_scalar(m, "variables_integer_nonbinary"), "Initial active non-binary integer variables."),
        ("Primary model continuous variables", "count", "Lower is better", lambda m: primary_scalar(m, "variables_continuous"), "Initial active continuous variables."),
        ("Primary model active constraints", "count", "Lower is better", lambda m: primary_scalar(m, "constraints_active"), "Before dynamic decomposition cuts."),
        ("Solver initial rows", "count", "Lower is better", lambda m: primary_solver(m, "Max initial_rows"), "From Gurobi logs."),
        ("Solver initial columns", "count", "Lower is better", lambda m: primary_solver(m, "Max initial_columns"), "From Gurobi logs."),
        ("Solver initial nonzeros", "count", "Lower is better", lambda m: primary_solver(m, "Max initial_nonzeros"), "From Gurobi logs."),
        ("Solver presolved rows", "count", "Lower is better", lambda m: primary_solver(m, "Max presolved_rows"), "Maximum observed primary-role presolved model."),
        ("Solver presolved columns", "count", "Lower is better", lambda m: primary_solver(m, "Max presolved_columns"), "Maximum observed primary-role presolved model."),
        ("Solver presolved nonzeros", "count", "Lower is better", lambda m: primary_solver(m, "Max presolved_nonzeros"), "Maximum observed primary-role presolved model."),
        ("Rows removed by presolve", "fraction", "Higher is better", lambda m: primary_solver(m, "Rows removed by presolve"), "Fraction of initial rows removed."),
        ("Columns removed by presolve", "fraction", "Higher is better", lambda m: primary_solver(m, "Columns removed by presolve"), "Fraction of initial columns removed."),
        ("Primary solver time", "s", "Lower is better", lambda m: primary_solver(m, "Total solver_seconds"), "Sum across primary-role solver calls."),
        ("Primary branch-and-bound nodes", "count", "Lower is better", lambda m: primary_solver(m, "Total nodes_explored"), "Sum across primary-role solves."),
        ("Primary solver work units", "work units", "Lower is better", lambda m: primary_solver(m, "Total work_units"), "Gurobi deterministic work measure."),
        ("Exact redirection variables simultaneously in primary model", "count", "Lower is better", lambda m: {"Monolithic": A * (C * C + 4), "LBBD": A}.get(m, math.nan), "Monolithic: type-pair + aggregate + activation + trips + tail; LBBD: continuous aggregate flow only."),
    ]
    efficiency_rows = []
    for metric, unit, direction, getter, notes in metric_specs:
        values = {method: _finite_float(getter(method)) for method in ("Monolithic", "LBBD")}
        mono = values["Monolithic"]
        value = values["LBBD"]
        row = {"Metric": metric, "Unit": unit, **values}
        row["LBBD difference vs Monolithic"] = (
            value - mono if math.isfinite(value) and math.isfinite(mono) else math.nan
        )
        if math.isfinite(value) and math.isfinite(mono) and abs(mono) > 1e-12:
            row["LBBD efficiency change vs Monolithic"] = (
                (mono - value) / abs(mono)
                if direction == "Lower is better" else (value - mono) / abs(mono)
            )
        else:
            row["LBBD efficiency change vs Monolithic"] = math.nan
        row["Preferred direction"] = direction
        row["Notes"] = notes
        efficiency_rows.append(row)

    return pd.DataFrame(l_rows), pd.DataFrame(efficiency_rows)

def _style_workbook(path: Path, requested: pd.DataFrame) -> None:
    wb = load_workbook(path)
    navy = "1F4E78"
    light_blue = "D9EAF7"
    pale = "EEF5FA"
    white = "FFFFFF"
    thin = Side(style="thin", color="B7C9D6")

    ws = wb["Requested comparison"]
    ws.freeze_panes = "B7"
    ws.auto_filter.ref = f"A6:E{ws.max_row}"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Monolithic–LBBD optimization comparison"
    ws["A1"].font = Font(size=16, bold=True, color=white)
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws.merge_cells("A1:E1")
    ws["A2"] = "Monolithic run"
    ws["A3"] = "LBBD run"
    for row in range(2, 4):
        ws.cell(row, 1).font = Font(bold=True)

    for col in range(1, 6):
        cell = ws.cell(6, col)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(bold=True, color=white)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)

    for idx, spec in enumerate(ROW_SPEC, start=7):
        _, _, indent, kind = spec
        metric_cell = ws.cell(idx, 1)
        metric_cell.alignment = Alignment(indent=indent, vertical="center")
        for col in range(1, 6):
            ws.cell(idx, col).border = Border(bottom=thin)
        if kind == "section":
            for col in range(1, 6):
                ws.cell(idx, col).fill = PatternFill("solid", fgColor=light_blue)
                ws.cell(idx, col).font = Font(bold=True, color=navy)
        elif indent:
            for col in range(1, 6):
                ws.cell(idx, col).fill = PatternFill("solid", fgColor=pale)

        if kind == "count":
            fmt = "#,##0"
        elif kind in {"currency", "number"}:
            fmt = "#,##0.00"
        elif kind == "percent":
            fmt = "0.0000%"
        else:
            fmt = "General"
        for col in (2, 3, 4):
            ws.cell(idx, col).number_format = fmt if kind != "section" else "General"
        ws.cell(idx, 5).number_format = "0.0000%"

    ws.column_dimensions["A"].width = 39
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 23
    ws.column_dimensions["E"].width = 25
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 21

    raw = wb["Raw metrics"]
    raw.freeze_panes = "B2"
    raw.auto_filter.ref = raw.dimensions
    raw.sheet_view.showGridLines = False
    for cell in raw[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(bold=True, color=white)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    raw.column_dimensions["A"].width = 44
    for col in range(2, raw.max_column + 1):
        raw.column_dimensions[get_column_letter(col)].width = 24
        for row in range(2, raw.max_row + 1):
            raw.cell(row, col).number_format = "0.000000"

    folders = wb["Run folders"]
    folders.sheet_view.showGridLines = False
    for cell in folders[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(bold=True, color=white)
    folders.column_dimensions["A"].width = 18
    folders.column_dimensions["B"].width = 110

    for sheet_name in [
        "Computational summary", "Build timing", "Redirection complexity",
        "Solver complexity", "Solver log detail", "Preprocessing scalars",
        "LBBD complexity", "Efficiency summary",
    ]:
        if sheet_name not in wb.sheetnames:
            continue
        extra = wb[sheet_name]
        extra.freeze_panes = "A2"
        extra.auto_filter.ref = extra.dimensions
        extra.sheet_view.showGridLines = False
        for cell in extra[1]:
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(bold=True, color=white)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(1, extra.max_column + 1):
            max_len = max(
                len(str(extra.cell(row, col).value or ""))
                for row in range(1, min(extra.max_row, 100) + 1)
            )
            extra.column_dimensions[get_column_letter(col)].width = min(42, max(12, max_len + 2))
        for row in range(2, extra.max_row + 1):
            for col in range(1, extra.max_column + 1):
                value = extra.cell(row, col).value
                if isinstance(value, (int, float)):
                    header = str(extra.cell(1, col).value or "").lower()
                    unit = str(extra.cell(row, 5).value or "").lower() if extra.max_column >= 5 else ""
                    if "efficiency change" in header or "fraction" in header or "gap" in header or "removed" in header or unit == "fraction":
                        extra.cell(row, col).number_format = "0.0000%"
                    elif unit == "count":
                        extra.cell(row, col).number_format = "#,##0"
                    else:
                        extra.cell(row, col).number_format = "#,##0.000"
        if sheet_name == "LBBD complexity":
            extra.column_dimensions["A"].width = 29
            extra.column_dimensions["B"].width = 39
            extra.column_dimensions["C"].width = 48
            extra.column_dimensions["D"].width = 22
            extra.column_dimensions["E"].width = 14
            extra.column_dimensions["F"].width = 55
            previous_component = None
            for row in range(2, extra.max_row + 1):
                component = extra.cell(row, 1).value
                if component != previous_component:
                    for col in range(1, extra.max_column + 1):
                        extra.cell(row, col).fill = PatternFill("solid", fgColor=light_blue)
                        extra.cell(row, col).font = Font(bold=True, color=navy)
                previous_component = component
        elif sheet_name == "Efficiency summary":
            extra.column_dimensions["A"].width = 43
            extra.column_dimensions["B"].width = 14
            for col in range(3, extra.max_column + 1):
                extra.column_dimensions[get_column_letter(col)].width = 24
            if extra.max_column >= 2:
                extra.column_dimensions[get_column_letter(extra.max_column)].width = 60
    wb.save(path)

def main() -> int:
    args = parse_args()
    root = Path(args.project_root).resolve()
    run_dirs = {
        "Monolithic": _resolve(root, args.monolithic_run),
        "LBBD": _resolve(root, args.lbbd_run),
    }
    capacity = _load_capacity_config(root)
    all_metrics = {
        method: _enrich(_read_summary(run_dir), run_dir, capacity)
        for method, run_dir in run_dirs.items()
    }

    shared_demand = next(
        (
            float(metrics.get("annual_public_demand_kWh"))
            for metrics in all_metrics.values()
            if math.isfinite(float(metrics.get("annual_public_demand_kWh", math.nan)))
        ),
        math.nan,
    )
    if math.isfinite(shared_demand):
        for metrics in all_metrics.values():
            demand = float(metrics.get("annual_public_demand_kWh", math.nan))
            if not math.isfinite(demand):
                metrics["annual_public_demand_kWh"] = shared_demand
            redirected = float(metrics.get("energy_redirected_kWh", 0.0) or 0.0)
            metrics["share_redirected_public_demand"] = (
                redirected / shared_demand if shared_demand > 0 else 0.0
            )

    requested = _requested_frame(all_metrics)
    raw = _raw_frame(all_metrics)
    computational_summary, build_timing, solver_complexity, solver_log_detail = _complexity_frames(run_dirs)
    dataset = _infer_dataset(run_dirs, args.dataset)
    scenario = _infer_scenario(run_dirs, args.scenario)
    vipv_scenario = _infer_vipv_scenario(run_dirs, args.vipv_scenario)

    preprocessing_frame = pd.DataFrame()
    redirection_complexity = pd.DataFrame()
    comparison_data = None
    if not args.skip_complexity_recompute:
        print(f"Recomputing exact preprocessing complexity for dataset={dataset}, scenario={scenario}...")
        try:
            comparison_data = _load_preprocessed_comparison_data(
                root, dataset, scenario, vipv_scenario
            )
            preprocessing_frame = pd.DataFrame([
                {"Metric": key, "Value": value}
                for key, value in preprocessing_scalars(comparison_data).items()
            ])
            redirection_complexity = slot_redirection_complexity(comparison_data)
        except Exception as exc:
            print(f"WARNING: Exact preprocessing complexity recomputation failed: {exc}")

    lbbd_complexity, efficiency_summary = _detailed_complexity_tables(
        comparison_data, run_dirs, solver_complexity
    )

    if args.out:
        output = _resolve(root, args.out)
    else:
        out_dir = root / "runs" / "comparisons"
        out_dir.mkdir(parents=True, exist_ok=True)
        output = out_dir / f"{datetime.now():%Y-%m-%d_%H%M%S}_monolithic_lbbd_comparison.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)

    visible_requested = requested.drop(columns=["MetricKey", "Indent", "Format"])
    run_folders = pd.DataFrame(
        [{"Method": method, "Run folder": str(path)} for method, path in run_dirs.items()]
    )
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        visible_requested.to_excel(
            writer, sheet_name="Requested comparison", index=False, startrow=5
        )
        raw.to_excel(writer, sheet_name="Raw metrics", index=False)
        run_folders.to_excel(writer, sheet_name="Run folders", index=False)
        computational_summary.to_excel(writer, sheet_name="Computational summary", index=False)
        build_timing.to_excel(writer, sheet_name="Build timing", index=False)
        redirection_complexity.to_excel(writer, sheet_name="Redirection complexity", index=False)
        solver_complexity.to_excel(writer, sheet_name="Solver complexity", index=False)
        solver_log_detail.to_excel(writer, sheet_name="Solver log detail", index=False)
        preprocessing_frame.to_excel(writer, sheet_name="Preprocessing scalars", index=False)
        lbbd_complexity.to_excel(writer, sheet_name="LBBD complexity", index=False)
        efficiency_summary.to_excel(writer, sheet_name="Efficiency summary", index=False)

    wb = load_workbook(output)
    ws = wb["Requested comparison"]
    ws["B2"] = str(run_dirs["Monolithic"])
    ws["B3"] = str(run_dirs["LBBD"])
    wb.save(output)
    _style_workbook(output, requested)

    csv_path = output.with_name(output.stem + "_requested_table.csv")
    visible_requested.to_csv(csv_path, index=False)
    print(f"Comparison workbook written to: {output}")
    print(f"Requested comparison CSV written to: {csv_path}")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
